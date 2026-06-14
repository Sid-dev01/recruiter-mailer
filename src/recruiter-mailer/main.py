from __future__ import annotations

import csv
import os
import re
import time
import smtplib
import mimetypes
from pathlib import Path
from typing import Iterable
from datetime import datetime
from dataclasses import dataclass
from email.message import EmailMessage

from jinja2 import Template
from dotenv import load_dotenv
from rich.console import Console
from openpyxl import load_workbook

console = Console()

# ============================================================
# CHANGE THESE VALUES ONLY
# ============================================================

# Change this file name when moving from test Excel to main Excel
EXCEL_FILE = Path("data/test_recruiters.xlsx")

# If your Excel has multiple sheets, write sheet name here.
# Keep it None to use the active/default sheet.
SHEET_NAME = None

# Change this range for each batch
FROM_SNO = 1
TO_SNO = 5

# Keep True while testing.
# Change to False only when you actually want to send emails.
DRY_RUN = True

# Optional safety limit.
# Keep None to preview/send the full selected SNo range.
LIMIT = None

# Delay between emails when DRY_RUN = False
DELAY_SECONDS = 5

# Email subject
SUBJECT_TEMPLATE = "Application for Full Stack Developer Position"

# Email body template
TEMPLATE_FILE = Path("templates/message.txt")

# Resume attachment
ATTACH_RESUME = True
RESUME_FILE = Path("resumes/Siddhartha_Gautam.pdf")

# Your profile links
GITHUB_LINK = "https://github.com/sid-dev01"
PORTFOLIO_LINK = "https://your-portfolio-link.com"
LINKEDIN_LINK = "https://www.linkedin.com/in/siddhartha-gautam-259525287/"

# Sending log
LOG_FILE = Path("sent_log.csv")

# ============================================================
# DO NOT CHANGE BELOW THIS LINE UNLESS YOU WANT TO EDIT LOGIC
# ============================================================

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

COLUMN_ALIASES = {
    "sno": ["sno", "s_no", "serial_no", "serial_number", "sr_no"],
    "name": ["name", "recruiter_name", "full_name"],
    "email": ["email", "email_id", "e_mail", "mail"],
    "title": ["title", "titale", "position", "designation", "role"],
    "company": ["company", "company_name", "organisation", "organization"],
}


@dataclass
class Recruiter:
    sno: int
    name: str
    email: str
    title: str
    company: str


@dataclass
class SMTPConfig:
    host: str
    port: int
    username: str
    password: str
    from_email: str
    from_name: str
    use_ssl: bool


def normalize_header(value: object) -> str:
    if value is None:
        return ""

    return (
        str(value)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace(".", "")
    )


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_sno(value: object) -> int | None:
    if value is None:
        return None

    try:
        return int(float(str(value).strip()))
    except ValueError:
        return None


def load_smtp_config() -> SMTPConfig:
    load_dotenv()

    required = [
        "SMTP_HOST",
        "SMTP_PORT",
        "SMTP_USERNAME",
        "SMTP_PASSWORD",
        "SMTP_FROM_EMAIL",
    ]

    missing = [key for key in required if not os.getenv(key)]

    if missing:
        raise RuntimeError(f"Missing environment variables: {', '.join(missing)}")

    return SMTPConfig(
        host=os.environ["SMTP_HOST"],
        port=int(os.environ["SMTP_PORT"]),
        username=os.environ["SMTP_USERNAME"],
        password=os.environ["SMTP_PASSWORD"],
        from_email=os.environ["SMTP_FROM_EMAIL"],
        from_name=os.getenv("SMTP_FROM_NAME", ""),
        use_ssl=os.getenv("SMTP_USE_SSL", "false").lower() in {"1", "true", "yes"},
    )


def find_column_indexes(headers: list[str]) -> dict[str, int]:
    indexes: dict[str, int] = {}

    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in headers:
                indexes[field] = headers.index(alias)
                break

    missing = [field for field in COLUMN_ALIASES if field not in indexes]

    if missing:
        raise ValueError(
            f"Missing required Excel columns: {', '.join(missing)}. "
            f"Found columns: {', '.join(headers)}"
        )

    return indexes


def read_recruiters(
    excel_path: Path,
    sheet_name: str | None,
    from_sno: int,
    to_sno: int,
) -> list[Recruiter]:
    workbook = load_workbook(excel_path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Excel file is empty.")

    headers = [normalize_header(value) for value in rows[0]]
    indexes = find_column_indexes(headers)

    recruiters: list[Recruiter] = []
    seen_emails: set[str] = set()

    for row_number, row in enumerate(rows[1:], start=2):
        if not row or all(value is None for value in row):
            continue

        sno = parse_sno(row[indexes["sno"]])

        if sno is None:
            console.print(f"[yellow]Skipping Excel row {row_number}: invalid SNo[/yellow]")
            continue

        if sno < from_sno or sno > to_sno:
            continue

        def get(field: str) -> str:
            index = indexes[field]
            return clean_cell(row[index]) if index < len(row) else ""

        recruiter = Recruiter(
            sno=sno,
            name=get("name"),
            email=get("email").lower(),
            title=get("title"),
            company=get("company"),
        )

        if not EMAIL_RE.match(recruiter.email):
            console.print(f"[yellow]Skipping SNo {recruiter.sno}: invalid email[/yellow]")
            continue

        if recruiter.email in seen_emails:
            console.print(
                f"[yellow]Skipping SNo {recruiter.sno}: duplicate email in selected range[/yellow]"
            )
            continue

        seen_emails.add(recruiter.email)
        recruiters.append(recruiter)

    recruiters.sort(key=lambda recruiter: recruiter.sno)

    return recruiters


def get_template_context(recruiter: Recruiter) -> dict[str, object]:
    return {
        "sno": recruiter.sno,
        "name": recruiter.name,
        "email": recruiter.email,
        "title": recruiter.title,
        "company": recruiter.company,
        "github_link": GITHUB_LINK,
        "portfolio_link": PORTFOLIO_LINK,
        "linkedin_link": LINKEDIN_LINK,
    }


def render_template(template_text: str, recruiter: Recruiter) -> str:
    return Template(template_text).render(**get_template_context(recruiter))


def attach_file(message: EmailMessage, file_path: Path) -> None:
    content_type, encoding = mimetypes.guess_type(file_path)

    if content_type is None or encoding is not None:
        content_type = "application/octet-stream"

    main_type, sub_type = content_type.split("/", 1)

    with file_path.open("rb") as file:
        message.add_attachment(
            file.read(),
            maintype=main_type,
            subtype=sub_type,
            filename=file_path.name,
        )


def build_email(
    smtp_config: SMTPConfig,
    recruiter: Recruiter,
    subject_template: str,
    body_template: str,
    resume_file: Path | None = None,
) -> EmailMessage:
    subject = render_template(subject_template, recruiter)
    body = render_template(body_template, recruiter)

    message = EmailMessage()

    if smtp_config.from_name:
        message["From"] = f"{smtp_config.from_name} <{smtp_config.from_email}>"
    else:
        message["From"] = smtp_config.from_email

    message["To"] = recruiter.email
    message["Subject"] = subject
    message.set_content(body)

    if resume_file is not None:
        attach_file(message, resume_file)

    return message


def append_log(
    log_file: Path,
    recruiter: Recruiter,
    status: str,
    error: str = "",
) -> None:
    file_exists = log_file.exists()

    with log_file.open("a", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "timestamp",
                "sno",
                "name",
                "email",
                "title",
                "company",
                "status",
                "error",
            ],
        )

        if not file_exists:
            writer.writeheader()

        writer.writerow(
            {
                "timestamp": datetime.now().isoformat(timespec="seconds"),
                "sno": recruiter.sno,
                "name": recruiter.name,
                "email": recruiter.email,
                "title": recruiter.title,
                "company": recruiter.company,
                "status": status,
                "error": error,
            }
        )


def send_messages(
    smtp_config: SMTPConfig,
    recruiters: Iterable[Recruiter],
    subject_template: str,
    body_template: str,
    delay_seconds: float,
    log_file: Path,
    resume_file: Path | None,
) -> None:
    smtp_class = smtplib.SMTP_SSL if smtp_config.use_ssl else smtplib.SMTP

    with smtp_class(smtp_config.host, smtp_config.port, timeout=30) as server:
        if not smtp_config.use_ssl:
            server.starttls()

        server.login(smtp_config.username, smtp_config.password)

        for recruiter in recruiters:
            try:
                message = build_email(
                    smtp_config=smtp_config,
                    recruiter=recruiter,
                    subject_template=subject_template,
                    body_template=body_template,
                    resume_file=resume_file,
                )

                server.send_message(message)

                console.print(
                    f"[green]Sent:[/green] SNo {recruiter.sno} -> {recruiter.email}"
                )

                append_log(
                    log_file=log_file,
                    recruiter=recruiter,
                    status="sent",
                )

                time.sleep(delay_seconds)

            except Exception as exc:
                console.print(
                    f"[red]Failed:[/red] SNo {recruiter.sno} -> {recruiter.email}"
                )

                append_log(
                    log_file=log_file,
                    recruiter=recruiter,
                    status="failed",
                    error=str(exc),
                )


def preview_messages(
    recruiters: list[Recruiter],
    subject_template: str,
    body_template: str,
    resume_file: Path | None,
) -> None:
    console.print("[bold yellow]DRY RUN MODE: no emails will be sent.[/bold yellow]")

    if resume_file is not None and resume_file.exists():
        console.print(f"[cyan]Resume attachment:[/cyan] {resume_file}")
    elif resume_file is not None:
        console.print(f"[yellow]Resume not found yet:[/yellow] {resume_file}")

    for recruiter in recruiters:
        subject_preview = render_template(subject_template, recruiter)
        body_preview = render_template(body_template, recruiter)

        console.rule(f"SNo {recruiter.sno} | {recruiter.email}")
        console.print(f"[bold]SNo:[/bold] {recruiter.sno}")
        console.print(f"[bold]To:[/bold] {recruiter.email}")
        console.print(f"[bold]Name:[/bold] {recruiter.name}")
        console.print(f"[bold]Title:[/bold] {recruiter.title}")
        console.print(f"[bold]Company:[/bold] {recruiter.company}")
        console.print(f"[bold]Subject:[/bold] {subject_preview}")

        if resume_file is not None:
            console.print(f"[bold]Attachment:[/bold] {resume_file}")

        console.print(body_preview)


def main() -> None:
    if FROM_SNO > TO_SNO:
        raise ValueError("FROM_SNO cannot be greater than TO_SNO.")

    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")

    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"Template file not found: {TEMPLATE_FILE}")

    selected_resume_file = RESUME_FILE if ATTACH_RESUME else None

    if selected_resume_file is not None and not selected_resume_file.exists():
        if DRY_RUN:
            console.print(
                f"[yellow]Warning: resume file not found yet: {selected_resume_file}[/yellow]"
            )
        else:
            raise FileNotFoundError(f"Resume file not found: {selected_resume_file}")

    template_text = TEMPLATE_FILE.read_text(encoding="utf-8")

    recruiters = read_recruiters(
        excel_path=EXCEL_FILE,
        sheet_name=SHEET_NAME,
        from_sno=FROM_SNO,
        to_sno=TO_SNO,
    )

    if LIMIT is not None:
        recruiters = recruiters[:LIMIT]

    if not recruiters:
        console.print("[red]No valid recruiters found for the selected SNo range.[/red]")
        return

    console.print(f"[cyan]Excel file:[/cyan] {EXCEL_FILE}")
    console.print(f"[cyan]Selected SNo range:[/cyan] {FROM_SNO} to {TO_SNO}")
    console.print(f"[cyan]Valid recruiters loaded:[/cyan] {len(recruiters)}")

    if DRY_RUN:
        preview_messages(
            recruiters=recruiters,
            subject_template=SUBJECT_TEMPLATE,
            body_template=template_text,
            resume_file=selected_resume_file,
        )
        return

    console.print("[bold red]LIVE MODE: emails will be sent now.[/bold red]")

    smtp_config = load_smtp_config()

    send_messages(
        smtp_config=smtp_config,
        recruiters=recruiters,
        subject_template=SUBJECT_TEMPLATE,
        body_template=template_text,
        delay_seconds=DELAY_SECONDS,
        log_file=LOG_FILE,
        resume_file=selected_resume_file,
    )


if __name__ == "__main__":
    main()